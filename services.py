#service.py
import sqlite3
import calendar
import os
import traceback
from openpyxl import Workbook
from datetime import datetime
from PySide2.QtCore import Qt, QSortFilterProxyModel, QRegularExpression
from PySide2.QtWidgets import QTableView, QMessageBox, QFileDialog, QLineEdit, QHBoxLayout  
from PySide2.QtGui import QStandardItemModel, QStandardItem
from openpyxl.styles import NamedStyle
from openpyxl.utils import get_column_letter

#setmodel table_view
def conectar():
    return sqlite3.connect("database_horas_extras.db")
#
def criar_tabela():
    if not os.path.exists("database_horas_extras.db"):
        meses = [
            "Janeiro", "Fevereiro", "Marco", "Abril", "Maio", "Junho",
            "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"
        ]
        
        with conectar() as conn:
            cursor = conn.cursor()
            ano_atual = datetime.now().year

            for mes in meses:
                numero_do_mes = meses.index(mes) + 1
                dias_no_mes = calendar.monthrange(ano_atual, numero_do_mes)[1]

                colunas = ['nome', 'supervisor', 'total_horas']
                for dia in range(1, dias_no_mes + 1):
                    data_formatada = f'{dia:02d}/{numero_do_mes:02d}/{ano_atual}'
                    colunas.append(data_formatada)

                valores_iniciais = []
                for col in colunas:
                    if col.lower() == 'nome':
                        valores_iniciais.append('"OperadorExemplo"')
                    elif col.lower() == 'supervisor':
                        valores_iniciais.append('"SupervisorExemplo"')
                    elif col.lower() == 'total de horas':
                        valores_iniciais.append('"00:00:00"')
                    else:
                        valores_iniciais.append('"00:00:00"')
                
                # Criar a tabela para o mês atual
                cursor.execute(f"""
                    CREATE TABLE IF NOT EXISTS "{mes}" (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        {', '.join([f'"{col}" text' for col in colunas])}
                    )
                """)


                # Insere a primeira linha 
                cursor.execute(f'''
                    INSERT INTO "{mes}" ({', '.join([f'"{col}"' for col in colunas])})
                    VALUES ({', '.join(valores_iniciais)})
                ''')
            
            conn.commit()
#
def carregar_supervisores(nome_tabela: str):
    conn = conectar()
    cursor = conn.cursor()

    cursor.execute(f"SELECT DISTINCT supervisor FROM {nome_tabela}")
    supervisores = cursor.fetchall()  # lista de tuplas [(nome1,), (nome2,), ...]

    cursor.close()
    conn.close()

    # Retorna apenas os nomes como uma lista simples
    lista_supervisores = [s[0] for s in supervisores if s[0] is not None]

    lista_supervisores.insert(0, "Todos")

    return lista_supervisores
#
def carregar_tabelaBancoDados(nome_tabela: str, supervisor: str, table_view: QTableView):
    #print("[DEBUG 1] este é o supervisor:", supervisor)
    #traceback.print_stack(limit=2)

    try:
        conn = conectar()
        cursor = conn.cursor()
        
        #print("este é o supervisor "+supervisor)
        if supervisor == "Todos":
            query = f"SELECT * FROM {nome_tabela}"
            cursor.execute(query)
        else:
            query = f"SELECT * FROM {nome_tabela} WHERE supervisor = ?"
            parametros = (supervisor,)
            cursor.execute(query, parametros)

        registros = cursor.fetchall()
        colunas = [description[0] for description in cursor.description]

        # Criar modelo de tabela
        modelo = QStandardItemModel()
        modelo.setColumnCount(len(colunas))
        modelo.setHorizontalHeaderLabels(colunas)

        indice_total = colunas.index("total_horas")

        for linha in registros:
            itens = []
            for i, campo in enumerate(linha):
                item = QStandardItem(str(campo))
                if i >= indice_total:
                    item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                itens.append(item)
            modelo.appendRow(itens)

        # Impedir a edição das colunas 'id', 'total_horas'
        for row in range(modelo.rowCount()):
            for col in range(modelo.columnCount()):
                nome_coluna = modelo.horizontalHeaderItem(col).text()
                if nome_coluna in ['id', 'total_horas']:
                    modelo.item(row, col).setFlags(modelo.item(row, col).flags() & ~Qt.ItemFlag.ItemIsEditable)
        
        proxy_model = QSortFilterProxyModel(table_view)
        proxy_model.setSourceModel(modelo)
        proxy_model.setFilterCaseSensitivity(Qt.CaseInsensitive)
        proxy_model.setSortCaseSensitivity(Qt.CaseInsensitive)

        table_view.setModel(proxy_model)
        table_view.setSortingEnabled(True)
        table_view.verticalHeader().setVisible(False)
        table_view.sortByColumn(1, Qt.AscendingOrder)  # por exemplo, ordena por "nome"

        # Criando layout_filtro_header
        layout_filtro_header = QHBoxLayout()  # Definindo o layout de filtros

        # Filtros:
        for i in range(modelo.columnCount()):
            filtro = QLineEdit()
            filtro.setPlaceholderText(f"Filtrar {colunas[i]}")
            filtro.setFixedHeight(25)
            layout_filtro_header.addWidget(filtro)

            def criar_filtro_funcao(indice):
                def filtro_funcao(texto):
                    proxy_model.setFilterKeyColumn(indice)
                    proxy_model.setFilterRegExp(QRegularExpression(texto, QRegularExpression.CaseInsensitiveOption))
                return filtro_funcao

            filtro.textChanged.connect(criar_filtro_funcao(i))

        table_view.sortByColumn(1, Qt.AscendingOrder)

        # Dicionário com larguras personalizadas para colunas
        larguras_personalizadas = {
            "id": 20,
            "nome": 200,
            "supervisor": 80,
            "total_horas": 100
        }

        # Ajustar largura das colunas específicas
        for i, nome_coluna in enumerate(colunas):
            if nome_coluna in larguras_personalizadas:
                table_view.setColumnWidth(i, larguras_personalizadas[nome_coluna])
            else:
                table_view.resizeColumnToContents(i)  # Tamanho automático para as demais

        def on_item_changed(item):
            try:
                modelo.itemChanged.disconnect()
                atualizar_celula_banco(item, supervisor, nome_tabela, table_view)
            finally:
                modelo.itemChanged.connect(on_item_changed)

        modelo.itemChanged.connect(on_item_changed)

    except Exception as e:
        QMessageBox.critical(table_view, "Erro", f"Erro ao carregar os dados: {str(e)}")

    finally:
        if conn:
            conn.close()

#
def atualizar_celula_banco(item: QStandardItem, supervisor: str, nome_tabela: str, table_view: QTableView):
    conn = None
    try:
        linha = item.row()
        coluna = item.column()
        novo_valor = item.text()
        model = item.model()

        nome_coluna = model.horizontalHeaderItem(coluna).text()
        id_item = model.item(linha, 0).text()

        # Ignorar colunas que não podem ser editadas
        if nome_coluna in ['id', 'total_horas']:
            return

        # Garantir formato HH:MM:SS
        if nome_coluna not in ['nome', 'supervisor']:
            if not novo_valor:
                return
            partes = novo_valor.strip().split(':')
            if len(partes) != 3:
                QMessageBox.warning(None, "Erro", "Formato de hora inválido! Use HH:MM:SS.")
                return

            try:
                horas, minutos, segundos = [p.zfill(2) for p in partes]
                novo_valor_formatado = f"{horas}:{minutos}:{segundos}"
                datetime.strptime(novo_valor_formatado, "%H:%M:%S")
            except ValueError:
                QMessageBox.warning(None, "Erro", "Formato de hora inválido! Use HH:MM:SS.")
                return
        else:
            # Para nome e supervisor, não altera o valor
            novo_valor_formatado = novo_valor.strip()

        # Atualizar a célula editada
        item.setText(novo_valor_formatado)

        # Encontrar o índice da coluna 'total_horas'
        indice_total = None
        for col in range(model.columnCount()):
            if model.horizontalHeaderItem(col).text() == "total_horas":
                indice_total = col
                break

        if indice_total is None:
            QMessageBox.critical(None, "Erro", "Coluna 'total_horas' não encontrada!")
            return

        # Recalcular total_horas da linha print
        total_segundos = 0
        for col in range(4, model.columnCount()):  # Começa a soma após a coluna 'nome'
            nome_col = model.horizontalHeaderItem(col).text()
            if nome_col == "total_horas":  # Ignorar a coluna 'total_horas' para soma
                continue

            valor = model.item(linha, col).text()
            try:
                h, m, s = map(int, valor.strip().split(":"))
                total_segundos += h * 3600 + m * 60 + s
            except TypeError:
                continue

        horas = total_segundos // 3600
        minutos = (total_segundos % 3600) // 60
        segundos = total_segundos % 60
        total_horas_str = f"{horas:02}:{minutos:02}:{segundos:02}"

        # Atualizar banco de dados
        conn = conectar()
        cursor = conn.cursor()

        if nome_coluna != "total_horas":
            cursor.execute(f'''
                UPDATE "{nome_tabela}"
                SET "{nome_coluna}" = ?, total_horas = ?
                WHERE id = ?
            ''', (novo_valor_formatado, total_horas_str, id_item))
        else:
            cursor.execute(f'''
                UPDATE "{nome_tabela}"
                SET total_horas = ?
                WHERE id = ?
            ''', (total_horas_str, id_item))

        conn.commit()
        carregar_tabelaBancoDados(nome_tabela, supervisor, table_view)

    except Exception as e:
        QMessageBox.critical(None, "Erro", f"Erro ao atualizar o banco de dados: {str(e)}")

    finally:
        if conn:
            conn.close()

#
def obter_ids_por_mes(mes):
    try:
        conn = conectar()
        cursor = conn.cursor()
        cursor.execute(f"SELECT id FROM {mes}")
        ids = [str(row[0]) for row in cursor.fetchall()]
        conn.close()
        return ids
    except Exception as e:
        print(f"Erro ao buscar IDs do mês {mes}: {e}")
        return []
    
def obter_dados_por_id(mes_selecionado, id_selecionado):
    try:
        conn = conectar()
        cursor = conn.cursor()
        
        # Buscar os dados do registro com o ID informado
        cursor.execute(f"SELECT * FROM {mes_selecionado} WHERE id = ?", (id_selecionado,))
        dados = cursor.fetchone()
        
        if dados:
            # Retorna os dados do registro, ou None se não encontrado
            colunas = [description[0] for description in cursor.description]
            return dict(zip(colunas, dados))
        else:
            return None
        
    except Exception as e:
        print(f"Erro ao obter dados do ID {id_selecionado}: {e}")
        return None
    finally:
        if conn:
            conn.close()
#
def exportar_tabela_para_xlsx(nome_tabela: str, parent_widget, caminho_salvar=None):
    try:
        conn = conectar()
        cursor = conn.cursor()
        cursor.execute(f'SELECT * FROM "{nome_tabela}"')
        registros = cursor.fetchall()
        colunas = [desc[0] for desc in cursor.description]

        # Se o caminho não for passado, abre o diálogo para salvar
        if not caminho_salvar:
            caminho_salvar, _ = QFileDialog.getSaveFileName(
                parent_widget, "Salvar como", f"{nome_tabela}.xlsx", "Arquivos Excel (*.xlsx)"
            )

        if not caminho_salvar:
            return  # Usuário cancelou

        # Criar planilha
        wb = Workbook()
        ws = wb.active
        ws.title = nome_tabela

        # Estilo para células de hora
        hora_style = NamedStyle(name="hora_style", number_format="HH:MM")

        # Escrever cabeçalho
        ws.append(colunas)

        # Escrever dados e formatar as colunas com hora
        for i, linha in enumerate(registros, start=2):  # Começa na linha 2 (linha 1 é o cabeçalho)
            for j, valor in enumerate(linha):
                # Se o valor for do tipo datetime ou hora, formate como hora
                if isinstance(valor, datetime):  # Caso seja datetime (para hora com data)
                    ws.cell(row=i, column=j + 1, value=valor).style = hora_style
                elif isinstance(valor, str) and ':' in valor:  # Caso seja string e tenha o formato de hora
                    try:
                        # Tentativa de converter para hora se estiver no formato "HH:MM:SS"
                        hora_valor = datetime.strptime(valor, '%H:%M:%S').time()
                        ws.cell(row=i, column=j + 1, value=hora_valor).style = hora_style
                    except ValueError:
                        ws.cell(row=i, column=j + 1, value=valor)  # Caso não seja hora válida
                else:
                    ws.cell(row=i, column=j + 1, value=valor)

        # Ajustar largura das colunas
        for col in range(1, len(colunas) + 1):
            max_length = 0
            column = get_column_letter(col)
            for row in ws.iter_rows(min_col=col, max_col=col):
                for cell in row:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except TypeError:
                        pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        # Salvar arquivo
        wb.save(caminho_salvar)

        QMessageBox.information(parent_widget, "Exportado", f"Tabela '{nome_tabela}' exportada com sucesso para:\n{caminho_salvar}")
    except Exception as e:
        QMessageBox.critical(parent_widget, "Erro", f"Erro ao exportar: {str(e)}")
    finally:
        conn.close()
#
def inserir_registro(mes_selecionado, dados, parent_widget=None):
    if not dados:
        QMessageBox.warning(parent_widget, "Aviso", "Por favor, insira os dados no campo de texto.")
        return

    meses = [
        "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
        "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"
    ]

    try:
        conn = conectar()
        cursor = conn.cursor()
        ano_atual = datetime.now().year

        numero_do_mes = meses.index(mes_selecionado) + 1
        dias_no_mes = calendar.monthrange(ano_atual, numero_do_mes)[1]

        colunas = ['nome', 'supervisor', 'total_horas'] + [
            f'{dia:02d}/{numero_do_mes:02d}/{ano_atual}' for dia in range(1, dias_no_mes + 1)
        ]

        linhas = dados.strip().split("\n")
        erros = []

        for linha in linhas:
            if not linha.strip():
                continue
            try:
                nome, supervisor = [parte.strip() for parte in linha.split(",")]
                valores = [nome, supervisor, "00:00:00"] + ["00:00:00"] * dias_no_mes

                colunas_sql = ', '.join([f'"{c}"' for c in colunas])
                valores_sql = ', '.join(['?'] * len(valores))
                comando_sql = f'''
                    INSERT INTO "{mes_selecionado}" ({colunas_sql})
                    VALUES ({valores_sql})
                '''
                cursor.execute(comando_sql, valores)

            except ValueError:
                erros.append(f"A linha '{linha}' não está no formato correto (esperado: nome, supervisor).")
            except sqlite3.Error as e:
                conn.rollback()
                QMessageBox.critical(parent_widget, "Erro no Banco de Dados", str(e))
                return

        conn.commit()

        if erros:
            QMessageBox.warning(None, "Linhas com Erros", "\n".join(erros))
        else:
            QMessageBox.information(None, "Sucesso", f"Todos os dados foram salvos na tabela '{mes_selecionado}'.")

    finally:
        if conn:
            conn.close()
#
def excluir_Registro(mes_selecionado, id_selecionado):
    conn = conectar()
    cursor = conn.cursor()
    cursor.execute(f"DELETE FROM {mes_selecionado} WHERE id = ?", (id_selecionado,))
    conn.commit()
    conn.close()
#
#def salvar_alteracoes():
    #asd
#